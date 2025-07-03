# Copyright (c) 2024, Frappe Technologies Pvt. Ltd. and Contributors
# License: GNU General Public License v3. See license.txt

import frappe
from frappe import _
from frappe.utils import (
    cint, flt, get_datetime, add_months, getdate, 
    today, format_datetime
)
from datetime import datetime
import json
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from io import BytesIO


def get_or_create_pricing_rule_for_amount(amount):
    """Get or create a pricing rule for the given gift card amount"""
    pricing_rule_name = f"Gift Card - {flt(amount, 2)}"
    
    existing_rule = frappe.db.exists("Pricing Rule", pricing_rule_name)
    if existing_rule:
        return existing_rule
    
    pricing_rule = frappe.new_doc("Pricing Rule")
    pricing_rule.update({
        "title": pricing_rule_name,
        "apply_on": "Transaction",
        "price_or_product_discount": "Price",
        "rate_or_discount": "Discount Amount",
        "discount_amount": flt(amount),
        "coupon_code_based": 1,
        "priority": 1,
        "company": frappe.defaults.get_user_default("Company"),
        "valid_from": today(),
        "valid_upto": "2999-12-31",  # Set a far future date like in the reference code
        "is_cumulative": 1,  # Changed to match reference code
        "apply_discount_on": "Grand Total",
        "apply_multiple_pricing_rules": 0,
        "selling": 1,  # Enable for selling transactions
        "buying": 0,   # Disable for buying transactions
        "margin_type": "Amount",  # Added from reference code
        "disable": 0  # Ensure it's enabled
    })
    pricing_rule.insert(ignore_permissions=True)
    
    return pricing_rule.name


def validate_gift_card_code(code):
    """Validate gift card code format and uniqueness"""
    if not code:
        return False, _("Gift card code cannot be empty")
    
    if len(code) < 4:
        return False, _("Gift card code must be at least 4 characters long")
    
    if frappe.db.exists("Coupon Code", {"coupon_code": code}):
        return False, _("Gift card code {0} already exists").format(code)
    
    return True, ""


def create_gift_card(code, amount, valid_until=None, customer_email=None, 
                    owner_email=None, name=None, default_customer=None):
    """Create a single gift card"""
    # Validate code
    is_valid, error_msg = validate_gift_card_code(code)
    if not is_valid:
        return {"success": False, "error": error_msg}
    
    try:        
        # Get or create pricing rule
        pricing_rule = get_or_create_pricing_rule_for_amount(amount)

        # Get webshop settings
        settings = frappe.get_single("Webshop Settings")
        
        # Calculate validity dates
        valid_from = today()
        if valid_until:
            if valid_until == "2999-12-31":
                valid_upto = valid_until  # Use as string for far future date
            else:
                valid_upto = getdate(valid_until)
        else:
            # Use default validity from settings or 12 months
            months = cint(settings.get("number_of_valid_months", 12))
            valid_upto = add_months(valid_from, months)
        
        # Create coupon code
        coupon = frappe.new_doc("Coupon Code")
        coupon.update({
            "coupon_name": name or code,
            "coupon_code": code,
            "coupon_type": "Gift Card",
            "pricing_rule": pricing_rule,
            "valid_from": getdate(valid_from),  # Ensure it's a date object
            "valid_upto": getdate(valid_upto),  # Ensure it's a date object
            "maximum_use": 9999,
            "used": 0,
            "description": f"Gift Card - {flt(amount, 2)} CHF"
        })
        
        # Set gift_card_amount separately to ensure it's not confused with date fields
        if hasattr(coupon, 'gift_card_amount'):
            coupon.gift_card_amount = flt(amount)
        
        # Set customer if email provided or use default
        customer = None
        if customer_email:
            customer = frappe.db.get_value("Customer", {"email_id": customer_email}, "name")
        
        # If no customer found and default customer provided, use it
        if not customer and default_customer:
            customer = default_customer
            
        if customer:
            coupon.customer = customer
        
        # Set owner if provided
        if owner_email and frappe.db.exists("User", owner_email):
            coupon.owner = owner_email
        
        coupon.insert(ignore_permissions=True)
        
        return {
            "success": True,
            "coupon_code": coupon.coupon_code,
            "amount": amount,
            "valid_until": format_datetime(valid_upto)
        }
        
    except Exception as e:
        frappe.log_error(title="Gift Card Creation Error", message=str(e))
        return {"success": False, "error": str(e)}


@frappe.whitelist()
def import_gift_cards_from_excel(file_content, dry_run=False, default_customer=None, default_validity_date=None):
    """Import gift cards from Excel file"""
    # Convert dry_run to boolean (it might come as string '0' or '1' from JavaScript)
    # When checkbox is checked, it sends 1, when unchecked it sends 0
    dry_run_value = frappe.utils.cint(dry_run)
   
    if not frappe.has_permission("Coupon Code", "create"):
        frappe.throw(_("You don't have permission to create gift cards"))
    
    try:
        # Decode base64 content
        import base64
        
        file_bytes = base64.b64decode(file_content)
        file_like = BytesIO(file_bytes)
        
        # Read Excel file using openpyxl
        workbook = openpyxl.load_workbook(file_like)
        sheet = workbook.active
        
        # Get headers from first row
        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value).strip().lower())
        
        # Validate required columns
        required_columns = ["code", "value"]
        missing_columns = [col for col in required_columns if col not in headers]
        if missing_columns:
            frappe.throw(_("Missing required columns: {0}").format(", ".join(missing_columns)))
        
        # Get column indices
        col_indices = {header: idx for idx, header in enumerate(headers)}
        
        results = {
            "success": [],
            "errors": [],
            "total": sheet.max_row - 1  # Exclude header row
        }
        
        # Process each row (skip header)
        for row_idx in range(2, sheet.max_row + 1):
            row_data = {}
            for header, col_idx in col_indices.items():
                cell_value = sheet.cell(row=row_idx, column=col_idx + 1).value
                row_data[header] = cell_value
            
            code = str(row_data.get("code", "")).strip() if row_data.get("code") else ""
            amount = flt(row_data.get("value", 0))
            name = str(row_data.get("name", "")).strip() if row_data.get("name") else None
            client_email = str(row_data.get("client_email", "")).strip() if row_data.get("client_email") else None
            owner_email = str(row_data.get("owner", "")).strip() if row_data.get("owner") else None
            valid_until = row_data.get("valid_until") if row_data.get("valid_until") else None
            
            # Validate amount
            if amount <= 0:
                results["errors"].append({
                    "row": row_idx,
                    "code": code,
                    "error": _("Amount must be greater than 0")
                })
                continue
            
            if dry_run_value:
                # Just validate without creating
                is_valid, error_msg = validate_gift_card_code(code)
                if is_valid:
                    results["success"].append({
                        "row": row_idx,
                        "code": code,
                        "amount": amount,
                        "message": _("Valid - ready to import")
                    })
                else:
                    results["errors"].append({
                        "row": row_idx,
                        "code": code,
                        "error": error_msg
                    })
            else:
                # Create the gift card                
                # Use default validity date if no date specified
                final_valid_until = valid_until or default_validity_date or "2999-12-31"
                
                # Use default customer if no customer email or customer not found
                final_customer_email = client_email
                if default_customer and not client_email:
                    # Get email of default customer
                    default_customer_email = frappe.db.get_value("Customer", default_customer, "email_id")
                    if default_customer_email:
                        final_customer_email = default_customer_email
                
                result = create_gift_card(
                    code=code,
                    amount=amount,
                    valid_until=final_valid_until,
                    customer_email=final_customer_email,
                    owner_email=owner_email,
                    name=name,
                    default_customer=default_customer
                )
                
                if result["success"]:
                    results["success"].append({
                        "row": row_idx,
                        "code": code,
                        "amount": amount,
                        "message": _("Created successfully")
                    })
                else:
                    results["errors"].append({
                        "row": row_idx,
                        "code": code,
                        "error": result["error"]
                    })
        
        if not dry_run_value and results["success"]:
            frappe.db.commit()
        
        return results
        
    except Exception as e:
        frappe.throw(_("Error importing gift cards: {0}").format(str(e)))


@frappe.whitelist()
def get_gift_card_import_template():
    """Generate and return a sample Excel template for gift card import"""
    # Create workbook and worksheet
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Gift Cards"
    
    # Define headers
    headers = ["name", "owner", "client_email", "code", "value", "valid_until"]
    
    # Sample data
    sample_data = [
        ["SUMMER-GIFT-001", "admin@example.com", "customer1@example.com", "SUMMER2024-001", 50.00, "2025-12-31"],
        ["BIRTHDAY-GIFT-002", "admin@example.com", "customer2@example.com", "BDAY2024-002", 100.00, "2025-06-30"],
        ["WELCOME-GIFT-003", "admin@example.com", "", "WELCOME2024-003", 25.00, ""]
    ]
    
    # Write headers with formatting
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col_idx, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = Alignment(horizontal='center')
    
    # Write sample data
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            sheet.cell(row=row_idx, column=col_idx, value=value)
    
    # Adjust column widths
    column_widths = [20, 25, 30, 20, 12, 15]
    for col_idx, width in enumerate(column_widths, 1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    
    # Add instructions sheet
    instructions_sheet = workbook.create_sheet("Instructions")
    instructions = [
        "Instructions for Gift Card Import:",
        "",
        "Required columns:",
        "- code: Unique gift card code",
        "- value: Gift card amount",
        "",
        "Optional columns:",
        "- name: Descriptive name for the gift card",
        "- owner: Email of the user who owns/created the card",
        "- client_email: Email of the customer who will use the card",
        "- valid_until: Expiry date (YYYY-MM-DD format). If empty, uses default validity period",
        "",
        "Note: Leave optional fields empty if not needed"
    ]
    
    for row_idx, instruction in enumerate(instructions, 1):
        instructions_sheet.cell(row=row_idx, column=1, value=instruction)
    
    instructions_sheet.column_dimensions['A'].width = 100
    
    # Save to BytesIO
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    
    # Return file info with base64 encoded content
    import base64
    return {
        "file_content": base64.b64encode(output.getvalue()).decode('utf-8'),
        "filename": "gift_cards_import_template.xlsx",
        "type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    }